import openpyxl
import numpy as np
import matplotlib.pyplot as plt
from mpl_toolkits.mplot3d import Axes3D

# 打开Excel文件
workbook = openpyxl.load_workbook("t_afternoon.xlsx")

# 选择第一个工作表
sheet = workbook.active

# 获取工作表的行数和列数
rows = sheet.max_row
columns = sheet.max_column

# 创建二维列表用于存储数据
a = [[None] * columns for _ in range(rows)]

# 遍历Excel表格中的所有单元格，并将值存储在a列表中
for index_row, row in enumerate(sheet.iter_rows(min_row=1, max_row=rows, min_col=1, max_col=columns), start=0):
    for index_column, cell in enumerate(row, start=0):
        if str(cell.value).find("i"):
            try:
                a[index_row][index_column] = complex(str(cell.value).replace("i", "j").replace(' ','')).real
            except ValueError:
                print(cell.value)
                input()
        else :
            a[index_row][index_column] = float(str(cell.value))
a = np.array(a)

max_index = np.unravel_index(np.argmax(a, axis=None), a.shape)
print("最大值的下标为:", max_index)

# 获取数组的维度
rows, cols = a.shape
x = np.arange(0, cols)
y = np.arange(0, rows)
x, y = np.meshgrid(x, y)

fig = plt.figure()
ax = fig.add_subplot(111, projection='3d')

scale = 1

colors = plt.cm.viridis((a - np.min(a)) / (np.max(a) - np.min(a)))

ax.plot_surface(x, -(-90 + y * 2), a/scale, alpha=0.6, facecolors=colors)
ax.scatter(max_index[1], -(-90 + max_index[0] * 2), a[max_index]/scale, s=100, marker='o', color='red', label='max:(' + str(max_index[0]) + ',' + str(-(-90 + max_index[1] * 2)) + ',' + str(round(a[max_index[0]][max_index[1]] / scale, 3)) + ')')
ax.set_xlabel(r'$\theta$')
ax.set_ylabel(r'$\mu$')
ax.set_zlabel(r'Time/h$')
ax.set_title(r"Afternoon Time-$\theta,\mu$")
plt.legend()
plt.show()